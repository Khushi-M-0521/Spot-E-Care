################################# $pot-E-Care #######################################
from tkinter import *
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.chart import PieChart, Reference

user_details = openpyxl.load_workbook("User_details.xlsx")
billBook = openpyxl.load_workbook('BillBook.xlsx')
usersheet = user_details.active
trendbook = openpyxl.load_workbook('TrendBook.xlsx')
trendsheet = trendbook.active
userCart = {}
item_price = {"Adobe Acrobat Pro DC for Teams": 2000, "Adobe Lightroom CC": 6500,
              "Adobe Photoshop CC for team": 7000, "Breevy": 2720, "CCleaner": 1600,
              "Dr.Fone-Data Recovery": 3000, "F.lux": 800, "Fences": 880, "Grammerly": 1000,
              "iA Writer": 2320, "Krisp": 7680, "Microsoft Office": 3000,
              "Microsoft Office 365": 2000, "Microsoft Teams": 1500, "Quick Heal": 824,
              "Revo Uninstaller Pro": 3120, "Stellar Data Recovery": 5000, "Tally": 7204,
              "Vyapar Billing Software": 724, "Wise Data Recovery": 6000, 'Zoho': 1610}
label = 0
global bs
bs = []
bs.append(billBook.sheetnames[0])
bs[0] = billBook.active
global username
username = ''
global totalPrice


def first():
    print("----------------  Welcome to $pot-E-Care !! -----------------")
    print("Enter \n1.Start Shopping\n2. User Credentials\n3. Exit")
    choice = int(input())
    print()
    if choice == 1:
        Login()
    elif choice == 2:
        till = False
        while till != True:
            list1 = checkPerson()
            flag = list1[0]
            person = list1[1]
            if flag == 1:
                print(f"USERNAME:      {usersheet['a' + str(person)].value}")
                print(f"PHONE NUMBER:  {usersheet['c' + str(person)].value}")
                print(f"EMAIL-ID:      {usersheet['d' + str(person)].value}")
                print(f"GENDER:        {usersheet['e' + str(person)].value}")
                print(f"Age:           {usersheet['f' + str(person)].value}\n")
                end()
                till = True
            else:
                print("Invalid Username orPassword.")
                print("Enter\n1. Exit\n2. Try Again")
                ch = int(input())
                print()
                if ch == 1:
                    end()
    else:
        end()


def user():
    root = Tk()
    myLabel = Label(root, text="Namaste User!☻")
    myLabel.pack()
    myLabel = Label(root, text="Sorry, we couldn't find you")
    myLabel.pack()

    myLabel = Label(root, text="Please Try again")
    myLabel.pack()
    root.geometry("300x150")
    root.attributes("-topmost", True)

    def Close():
        root.destroy()
        # Button for closing

    try_button = Button(root, text="Try Again", command=Close)
    try_button.pack(pady=20)
    root.mainloop()


def crossCheck(message, dtype, condition, size):
    till = False
    while till == False:
        print("Enter ", message, " (", condition, size, " characters): ", end="")
        if dtype == 'i':
            inp = int(input())
            print()
            length = len(str(inp))
        elif dtype == 's':
            inp = input()
            print()
            length = len(inp)
        elif dtype == 'e':
            inp = input()
            print()
            length = len(inp)
            if not (inp[-10:len(inp)] == '@gmail.com' or inp[-12:len(inp)] == "@outlook.com"):
                print("Invalid Email-id")
                continue

        if condition == 'minimum':
            if length >= size:
                till = True
                return inp
            else:
                print(f"Invalid {message}. Enter {condition} of {size} character")
        elif condition == 'maximum':
            if length == size:
                till = True
                return inp
            else:
                print(f"Invalid {message}. Enter {condition} of {size} character")
        elif condition == 'M/F/O':
            if inp == 'M' or inp == 'F' or inp == 'O':
                till = True
                return inp
            else:
                print("Invalid Gender.")


def signUp():
    x = len(usersheet['a'])
    global username

    username = crossCheck("Username", 's', 'minimum', 6)
    usersheet['a' + str(x + 1)] = username

    password = crossCheck("Password", 's', 'minimum', 6)
    usersheet['b' + str(x + 1)] = password

    phone = crossCheck("Phone Number", 'i', 'maximum', 10)
    usersheet['c' + str(x + 1)] = phone

    email_id = crossCheck("Email-id", 'e', 'minimum', 11)
    usersheet['d' + str(x + 1)] = email_id

    gender = crossCheck("Gender", 's', 'M/F/O', 1)
    usersheet['e' + str(x + 1)] = gender

    age = crossCheck("Age", 'i', 'maximum', 2)
    usersheet['f' + str(x + 1)] = age

    user_details.save("User_details.xlsx")
    print("Enter \n1. Login\n2.Exit")
    choice = int(input())
    if choice == 1:
        Login()
    else:
        end()


def checkPerson():
    x = len(usersheet['a'])
    flag = 0
    row = 0
    global username
    username = input("Enter Username: ")
    password = input("Enter Password: ")
    print()
    for person in range(2, x + 1):
        if username == usersheet['a' + str(person)].value:
            if password == usersheet['b' + str(person)].value:
                flag = 1
                row = person
    return [flag, row]


def Login():
    list1 = checkPerson()
    flag = list1[0]
    if flag == 1:
        print("Login Successfull\n")
        cart()
    else:
        print("꒰⍨꒱")
        user()
        print("1 Login \n2. Sign UP \n3Exit")
        choice = int(input())
        print()
        if choice == 1:
            Login()
        elif choice == 2:
            signUp()
        else:
            end()


def itemsDetails():
    print("Code\tPACKAGES\t\t\t\tPRICE(per year)")
    len_packageName = max(len(name) for name in list(item_price.keys()))
    for package in range(len(item_price)):
        print("{0}\t{1:{2}}\t\t Rs.{3}".format(package + 1,
                                               list(item_price.keys())[package],
                                               len_packageName,
                                               list(item_price.values())[package]))
    print()


def cart():
    till = False
    while till != True:
        print("Enter\n1. Add Package(s)\n2. View Cart")
        choice = int(input())
        print()
        if choice == 1:
            itemsDetails()
            pack = int(input("Enter the Package code: "))
            while till != True:
                duration = int(input("Enter validity(max 3years): "))
                if duration > 3:
                    print("Validity is maximum for 3 years")
                else:
                    print()
                    till = True
            till = False
            userCart.update({list(item_price.keys())[pack - 1]: [list(item_price.values())[pack - 1], duration,
                                                                 list(item_price.values())[pack - 1] * duration]})
        else:
            if len(userCart) == 0:
                print("NO PACKAGES IN CART.")
                print("Enter\n1. Exit\n2. Continue")
                ch = int(input())
                print()
                if ch == 1:
                    end()
            else:
                print(f"You have {len(userCart)} packages in cart.")
                len_packageName = max(len(name) for name in list(userCart.keys()))
                print("{0}\t{1:{2}}\t{3}\t\t{4}".format("CODE",
                                                        "tPACKAGE",
                                                        len_packageName,
                                                        "PRICE",
                                                        "DURATION"))
                for elem in range(len(userCart)):
                    print("{0}\t{1:{2}}\tRs.{3}\t\t{4}".format(elem + 1,
                                                               list(userCart.keys())[elem],
                                                               len_packageName,
                                                               list(userCart.values())[elem][0],
                                                               list(userCart.values())[elem][1]))
                flag = 0
                print("\nEnter\n1. Bill\n2. Discard Package\n3. Continue")
                ch = int(input())
                print()
                if ch == 1:
                    flag = 1
                    till = True
                elif ch == 2:
                    delete()
    if flag == 1:
        bill()


def delete():
    code = int(input("Enter the Package Code to be discarded: "))
    trash = userCart.pop(list(userCart.keys())[code - 1])
    print("Dicarded successfully.")


def designTrend():
    for row in range(2, len(item_price) + 2):
        trendsheet['a' + str(row)] = row - 1
        trendsheet['b' + str(row)] = list(item_price.keys())[row - 2]
        number = trendsheet['c' + str(row)].value
        if number == None:
            trendsheet['c' + str(row)] = 0


def designChart():
    trendsheet = trendbook.active
    if len(trendsheet._charts) == 1:
        del trendsheet._charts[0]
    chart = PieChart()
    data = Reference(worksheet=trendsheet, min_col=3, min_row=1, max_row=len(item_price) + 1)
    labels = Reference(worksheet=trendsheet, min_col=2, min_row=2, max_row=len(item_price) + 1)
    chart.title = 'Customers per Package(in %)'
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.height = 16.20
    chart.width = 20
    trendsheet.add_chart(chart, 'F1')
    trendbook.save("TrendBook.xlsx")


def designBill(sh_num):
    global totalPrice
    totalPrice = 0
    page = bs[sh_num - len(billBook.sheetnames)]
    page.merge_cells('a1:e2')
    page['a1'].font = Font(size=16, bold=True, underline='double')
    page['a1'].alignment = Alignment(horizontal='center', vertical='center')
    page['a1'] = "BILL NO. " + str(sh_num + 1)
    page['a3'].font = Font(bold=True, underline='single')
    page['a3'].alignment = Alignment(horizontal='center', vertical='center')
    page['a3'] = "NAME: "
    page['b3'] = username
    page['c3'].font = Font(bold=True, underline='single')
    page['c3'].alignment = Alignment(horizontal='center', vertical='center')
    page['c3'] = 'EMAIL-ID:'
    x = len(usersheet['a'])
    for person in range(2, x + 1):
        if username == usersheet['a' + str(person)].value:
            page['d3'] = usersheet['d' + str(person)].value
    for row in 'abcde':
        page[str(row) + '4'].font = Font(bold=True, underline='single')
        page[str(row) + '4'].alignment = Alignment(horizontal='center', vertical='center')
    page['a4'] = "SL. NO."
    page['b4'] = "PACKAGE"
    page.column_dimensions['b'].width = 40
    page['c4'] = "DURATION(in yrs)"
    page.column_dimensions['c'].width = 20
    page['d4'] = "PACKAGE PRICE(Rs. per year)"
    page.column_dimensions['d'].width = 30
    page['e4'] = "PACKAGE COST(Rs.)"
    page.column_dimensions['e'].width = 20
    for row in range(len(userCart)):
        page['a' + str(row + 5)] = row + 1
        page['b' + str(row + 5)] = list(userCart.keys())[row]
        page['c' + str(row + 5)] = list(userCart.values())[row][1]
        page['D' + str(row + 5)] = list(userCart.values())[row][0]
        page['E' + str(row + 5)] = list(userCart.values())[row][2]
        totalPrice += list(userCart.values())[row][2]

        for item in range(2, len(item_price) + 2):
            if list(userCart.keys())[row] == trendsheet['B' + str(item)].value:
                trendsheet['C' + str(item)] = trendsheet['C' + str(item)].value + 1

    designChart()
    page['d' + str(len(userCart) + 5)].font = Font(bold=True, underline='single')
    page['d' + str(len(userCart) + 5)].alignment = Alignment(horizontal='center', vertical='center')
    page['d' + str(len(userCart) + 5)] = 'TOTAL PRICE(Rs.): '
    page['e' + str(len(userCart) + 5)] = totalPrice
    billBook.save('BillBook.xlsx')


def bill():
    global totalPrice
    designTrend()
    sheets = len(billBook.sheetnames)
    if sheets == 1:
        if bs[0].title != "BILL1":  # not filled
            bs[0].title = 'BILL1'
            designBill(0)
        else:  # filled
            bs.append(billBook.create_sheet("BILL2"))
            designBill(sheets)
    else:
        bs.append(billBook.create_sheet("BILL" + str(sheets + 1)))
        designBill(sheets)
    designChart()
    print(f'YOUR BILL: RS.{totalPrice}')
    end()


def end():
    print("------------------ Thank you ------------------------")
    print("---------------- Shop Again !! ----------------------")
    exit


first()
